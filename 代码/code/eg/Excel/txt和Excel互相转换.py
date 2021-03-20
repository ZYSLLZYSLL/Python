#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2021/1/19 10:34
# @Author : ZY
# @File : txt和Excel互相转换.py.py
# @Project : code

# 新建a.txt并写入内容
f = open('a.txt', 'w', encoding='utf-8')
f.write('1\na,b,c,d\n123,4,5,6\n中,国,富,强,123')
f.close()

# txt转Excel

# from openpyxl import load_workbook
# from openpyxl import Workbook
# j = 1  # 行
# k = 1  # 列
# a = ''
# wb = Workbook()
# # print(wb.sheetnames)  # Sheet
# f = open('a.txt', 'r', encoding='utf-8')
# fr = f.read()
#
# # fr = fr.split(',')
# ws = wb['Sheet']
# for i in fr:
#     if i == '\n':
#         ws.cell(j, k, a)
#         j += 1
#         k = 1
#         a = ''
#     else:
#         if i == ',':
#             ws.cell(j, k, a)
#             a = ''
#             k += 1
#         else:
#             a += i
# ws.cell(j, k, a)
# # print(list(fr))
#
# wb.save('a.xlsx')
# f.close()



# txt转Excel

from openpyxl import load_workbook
from openpyxl import Workbook

# k = 1  # 列
#
# wb = Workbook()
# # print(wb.sheetnames)  # Sheet
# ws = wb['Sheet']
#
# f = open('a.txt', 'r', encoding='utf-8')
# fr = f.read().split('\n')
# print(fr)
# for i in range(len(fr)):
#     k = 1
#     a = ''
#     for j in fr[i]:
#         if j == ',':
#             ws.cell(i + 1, k, a)
#             a = ''
#             k += 1
#         else:
#             a += j
#     ws.cell(i + 1, k, a)

# wb.save('a.xlsx')
# f.close()
