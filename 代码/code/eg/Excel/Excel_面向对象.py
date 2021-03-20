#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2021/1/19 14:05
# @Author : ZY
# @File : Excel_面向对象.py
# @Project : code

from openpyxl import load_workbook
from openpyxl import Workbook


class ExcelMutualTxt():
    def __init__(self):
        self.lie = 1
        self.temp_data = ''
        self.data_list = []

    def run(self):
        self.show_menu()
        while True:
            menu_num = int(input('请输入您需要的功能序号:'))

            if menu_num == 1:
                self.e_zhaun_t()
            elif menu_num == 2:
                self.t_zhaun_e()
            elif menu_num == 3:
                break

    @staticmethod
    def show_menu():
        print('请选择如下功能——————————————————————')
        print('1:Excel转Txt')
        print('2:Txt转Excel')
        print('3:退出系统')

    def e_zhaun_t(self):

        wb = load_workbook('a.xlsx')
        ws = wb['Sheet']

        f = open('a.txt', 'w+', encoding='utf-8')
        # print(w s.rows)  # 表内容的地址 <generator object Worksheet._cells_by_row at 0x000001D5425917C8>
        for i in ws.rows:
            #     # print(i)  # 元组 (<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>, <Cell 'Sheet'.D1>, <Cell 'Sheet'.E1>)
            #     # print(list(i))  # 列表 [<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>, <Cell 'Sheet'.D1>, <Cell 'Sheet'.E1>]
            self.data_list.append(list(i))

        for i in self.data_list:
            a = 0
            for j in i:
                if j.value == None:
                    f.write('')
                else:
                    if a == 1:
                        f.write(',')  # 先判断这一行后面还有没有数据，如果有数据才写逗号，如果没有数据就不写逗号
                        a = 0
                    f.write(f'{j.value}')
                    a = 1
                # if j.value == None:
                #     f.write('\t')
                # elif i[-1].value == None and j.value == i[-2].value:
                #     f.write(f'{j.value}')
                # elif i[-1].value != None and j.value == i[-1].value:
                #     f.write(f'{j.value}')
                # else:
                #     f.write(f'{j.value},')

            #  只有最后一行不换行，其他都换行
            if i != self.data_list[-1]:
                f.write('\n')
        # for i in f.readlines():
        #
        #     f.write(f'{f.readline()[:-1]}')

        # print(self.data_list)
        print('Excel转TxT转换完成')

    def t_zhaun_e(self):
        wb = Workbook()
        # print(wb.sheetnames)  # Sheet
        ws = wb['Sheet']

        f = open('a.txt', 'r', encoding='utf-8')
        fr = f.read().split('\n')  # 分割字符串

        for i in range(len(fr)):
            self.lie = 1
            self.temp_data = ''
            for j in fr[i]:
                if j == ',' or j == '，':  # 如果是逗号，证明上一个字符串已经传完了，写入excel
                    ws.cell(i + 1, self.lie, self.temp_data)
                    self.temp_data = ''
                    self.lie += 1
                else:  # 如果不是逗号，证明上一个字符串还没有传完
                    self.temp_data += j
            ws.cell(i + 1, self.lie, self.temp_data)

        wb.save('a.xlsx')
        f.close()
        print('TxT转Excel转换完成')


aa = ExcelMutualTxt()
aa.run()
