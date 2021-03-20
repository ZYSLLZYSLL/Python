#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2021/1/20 16:34
# @Author : ZY
# @File : Demo23_Excel内容到MySQL.py
# @Project : code

import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('爬虫.xlsx')
# print(wb.sheetnames)  # Sheet
sw1 = wb['Sheet']

root = pymysql.connect(host='192.168.2.118', user='root', password='123456')
cursor = root.cursor()

cursor.execute('create database if not exists Etomysql;')
cursor.execute('use Etomysql')
cursor.execute('create table if not exists web (`名字` char(100),`网址` char(200))')

for i in range(len(list(sw1.columns)[0])):
    # for j in list(sw1.rows[i]):
    cursor.execute(f"insert into web values ('{list(sw1.rows)[i][0].value}','{list(sw1.rows)[i][1].value}')")
# a = cursor.fetchall()
# print(a)

root.commit()
root.close()



