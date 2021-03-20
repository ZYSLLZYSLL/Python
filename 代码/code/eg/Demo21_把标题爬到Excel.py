#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2021/1/20 9:08
# @Author : ZY
# @File : Demo21_把标题爬到Excel.py
# @Project : code

import re
import requests
from openpyxl import Workbook

hang = 1
lie = 1

html = requests.get('https://www.ivsky.com/tupian/jiaotongyunshu/')
a1 = html.content.decode('utf-8')  # 找到网页的源代码
a2 = re.findall('ul class="ali"(.*?)</ul>', a1)  # 找到图片链接和名字的一段源代码
jpg = re.findall('<img src="(.*?)" ', a2[0])  # 找到图片的网址
title = re.findall('alt="(.*?)">', a2[0])  # 找到图片的名字

wb = Workbook()
# print(wb.sheetnames)  # 找到标签的名字Sheet
ws1 = wb['Sheet']
# for i in range(len(title)):  # 成行显示
#     if hang == 1:
#         ws1.cell(hang, lie, title[i])
#         hang =2
#     if hang == 2:
#         ws1.cell(hang, lie, "https:" + jpg[i])
#         hang = 1
#     lie += 1

for i in range(len(title)):  # 成列显示
    if lie == 1:
        ws1.cell(hang, lie, title[i])
        lie = 2
    if lie == 2:
        ws1.cell(hang, lie, "https:" + jpg[i])
        lie = 1
    hang += 1
wb.save('爬虫.xlsx')



# import openpyxl
# import requests
# import re
#
# def get_title():
#     r"""Sends a GET request.
#
#        :param url: url地址
#        :param params: (optional) Dictionary, ?键=值  {键：值}
#        :param \*\*kwargs: Optional arguments that ``request`` takes.
#        :return: :class:`Response <Response>` object  响应对象
#        :rtype: requests.Response
#        """
#     html = requests.get("https://www.ivsky.com/tupian/jiaotongyunshu/")
#     # content是以二进制的方式展示  encode 编码 decode 解码
#     result = html.content.decode("utf-8")
#     c = re.findall("<ul(.*?)</ul>",result)  # ['ul','ul','ul']
#     # print(c[2])
#     jpg= re.findall('img src="(.*?)"',c[2])
#     title = re.findall('alt="(.*?)">',c[2])
#     return title,jpg
#
# def write_excel(gt):
#     # 1. 创建工作薄
#     wb = openpyxl.Workbook()
#     # 2. 获取标签页
#     ws = wb["Sheet"]
#     # 3. 写入数据
#     for z in range(len(gt[0])):
#         ws.cell(z+1,1,gt[0][z])
#         ws.cell(z+1,2,gt[1][z])
#     wb.save("tupian.xlsx")
#
# write_excel(get_title())

