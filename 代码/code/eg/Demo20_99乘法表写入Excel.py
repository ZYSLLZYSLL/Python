#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2021/1/18 19:35
# @Author : ZY
# @File : Demo20_99乘法表写入Excel.py
# @Project : code

from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet('99乘法表')
for i in range(1, 10):
    for j in range(1, i + 1):
        ws.cell(i, j, f'{j}*{i}={i * j}')
wb.save('b.xlsx')
