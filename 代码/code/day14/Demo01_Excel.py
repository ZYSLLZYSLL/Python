#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2021/1/19 9:09
# @Author : ZY
# @File : Demo01_Excel.py
# @Project : code

from openpyxl import Workbook  # 加载工作簿
from openpyxl import load_workbook  # 加载工作簿

"""
    写
"""

# # 1，打开工作簿  write_only =False 只 写 模式
# wb = Workbook()  # 打开或创建,不能打开已经有的

# # 2，创建标签页
# ws = wb.create_sheet('bowen')

# # 3，写入单元格内容  def cell(self, row行, column列, value=None内容):
# ws.cell(1, 1, '周宇')

# # 4，保存文件
# wb.save('a.xlsx')

"""
    读
"""

# 1，打开工作簿
# wb = Workbook()  # 打开或创建

# 1，打开工作簿  write_only =False 默认可读可写
wb = load_workbook(r"c:\Users\ZYSLL\Desktop\www.xlsx")

# 2，创建标签页
ws = wb.create_sheet('bowen')

# a，获取标签页名称 返回一个列表
print(wb.sheetnames)  # ['Sheet', 'bowen']

# b，通过名字进入标签页
ws1 = wb['Sheet1']
ws1.cell(1, 1, "周宇")  # 向‘Sheet’标签页里写内容
ws1.cell(2, 1, "周宇1")  # 向‘Sheet’标签页里写内容
ws1.cell(1, 2, "周宇")  # 向‘Sheet’标签页里写内容

# d，追加到末尾（最后一行）
ws1.append([1, 2, 3, 4, 5])

# c，读行 row(行)
print(list(ws1.rows)[0][0].value)  # 这样出来才是数据
print(ws1.rows)  # <generator object Worksheet._cells_by_row at 0x00000204319F38C8>
print(list(ws1.rows))  # [(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>)]

# for i in ws1.rows:
#     print(i[0].value, i[1].value)

# 读列 columns(列)
print(list(ws1.columns))
# for i in ws1.columns:
#     print(i[0].value, i[1].value)

# d，有范围的读行
# 从第三行开始读到第十行，读三列
# for i in ws1.iter_rows(min_row=3, max_col=3, max_row=10):
#     print(i)

# e，读某个单元格的内容
# print(ws1['A1'].value)
# print(ws1['A1':'B2'])  # 一个范围内

# 3，写入单元格内容  def cell(self, row行, column列, value=None内容):
ws.cell(1, 1, '周宇')

# 4，保存文件
# wb.save('a.xlsx')
wb.save(r'c:\Users\ZYSLL\Desktop\www.xlsx')