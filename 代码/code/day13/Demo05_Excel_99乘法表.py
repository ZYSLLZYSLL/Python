#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2021/1/18 17:43
# @Author : ZY
# @File : Demo05_Excel_99乘法表.py
# @Project : code
from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet('99乘法表')
for i in range(1, 10):
    for j in range(1, i + 1):
        ws.cell(i, j, f'{j}*{i}={i * j}')
wb.save('b.xlsx')
