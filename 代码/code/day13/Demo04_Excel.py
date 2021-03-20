#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2021/1/18 17:13
# @Author : ZY
# @File : Demo04_Excel.py
# @Project : code

from openpyxl import load_workbook  # 加载工作簿 只能打开一个存在的
from openpyxl import Workbook  # 加载工作簿，只能打开一个不存在的

# filename string 文件名或路径
# read_only(只读模式) bool True 开始

# 1，打开工作簿  write_only =False 只写模式
wb = Workbook()  # 打开或创建
# 2，创建标签页
ws = wb.create_sheet('bowen')
# 3，写入单元格内容  def cell(self, row行, column列, value=None内容):
ws.cell(1, 1, '周宇')
# 4，保存文件
wb.save('a.xlsx')
